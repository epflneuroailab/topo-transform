"""Export the Nature Communications Source Data workbook for TopoTransform main figures.

One worksheet per *graph* panel (schematics and cortical-sheet maps are excluded by
policy; their underlying arrays live in the code/data release cited in Data
Availability). Every panel reuses the exact cached data-producing functions behind
the published figures, so the exported numbers match the plots. Bar-graph panels
include the individual data points (per-seed model instances / per-subject humans).

Run on the yt-active pod inside the dyviz env:
    python -m scripts.export_source_data              # cheap cached panels only
    python -m scripts.export_source_data --heavy      # + Fig 2b/2c (feature load + inference)
"""
import argparse
import io
import os
import traceback
from pathlib import Path

import numpy as np

import config
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PLOTS_DIR = config.PLOTS_DIR
DEBUG_DIR = config.DEBUG_DIR

# Store source-data values at high precision. Builders pass an explicit ndigits
# (e.g. round(x, 5)) for readability, but we override it here so nothing is
# under-rounded in the delivered file.
_builtin_round = round


def round(value, ndigits=None):  # noqa: A001 - intentional shadow to raise stored precision
    return _builtin_round(float(value), 9)
OUT_DIR = config.HOME_DIR / "source_data"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / "Source Data.xlsx"


# --------------------------------------------------------------------------- #
# cache guards: never trigger a recompute; skip checkpoints without a cache    #
# --------------------------------------------------------------------------- #
import hashlib


def _cache_exists(name):
    return (DEBUG_DIR / f"{name}.pkl").exists()


def _decode_cache_name(ckpt):
    from scripts.plot_localizer_decode import ROIS, _CLUSTER_K

    rc = hashlib.md5("_".join(sorted(ROIS)).encode()).hexdigest()[:8]
    kc = hashlib.md5("_".join(f"{k}:{v}" for k, v in sorted(_CLUSTER_K.items())).encode()).hexdigest()[:8]
    return f"localizer_decode_clustered_splits1_{ckpt}_rois{rc}_k{kc}_fwhm2.0_res1.0"


def _localizers_cache_name(ckpt):
    from validate.floc.registry import LOCALIZER_DATASETS

    h = hashlib.md5("_".join(sorted(LOCALIZER_DATASETS)).encode()).hexdigest()[:8]
    return f"localizers_{ckpt}_{h}_2.0_1.0"


def _smoothness_cache_name(ckpt, dataset, fwhm):
    return f"smoothness_{ckpt}_{dataset}_{fwhm}_1.0"


def _cached_ckpts(ckpts, kind):
    """Return (kept, dropped) checkpoints, keeping only those with an existing cache."""
    namer = {"decode": _decode_cache_name, "localizers": _localizers_cache_name}[kind]
    kept = [c for c in ckpts if _cache_exists(namer(c))]
    dropped = [c for c in ckpts if c not in kept]
    return kept, dropped

TITLE_FONT = Font(bold=True, size=13)
DESC_FONT = Font(italic=True, size=9, color="444444")
NOTE_FONT = Font(italic=True, size=9, color="8A5000")
BLOCK_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
WRAP = Alignment(wrap_text=True, vertical="top")


# --------------------------------------------------------------------------- #
# thumbnail helpers                                                           #
# --------------------------------------------------------------------------- #
def _svg_to_png(svg_path, png_path, scale=1.4):
    try:
        import cairosvg

        cairosvg.svg2png(url=str(svg_path), write_to=str(png_path), scale=scale)
        return True
    except Exception as exc:  # pragma: no cover - best effort
        print(f"    [thumb] cairosvg failed for {svg_path}: {exc}")
        return False


def _thumbnail_png(candidates, key):
    for rel in candidates:
        svg = PLOTS_DIR / rel
        if svg.exists():
            png = OUT_DIR / f"_thumb_{key}.png"
            if _svg_to_png(svg, png):
                return png
    return None


# --------------------------------------------------------------------------- #
# worksheet rendering                                                         #
# --------------------------------------------------------------------------- #
def _write_panel(ws, panel):
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 15

    r = 1
    ws.cell(r, 1, panel["title"]).font = TITLE_FONT
    r += 1
    c = ws.cell(r, 1, panel.get("description", ""))
    c.font = DESC_FONT
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 42
    r += 2

    for note in panel.get("notes", []):
        c = ws.cell(r, 1, note)
        c.font = NOTE_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1
    if panel.get("notes"):
        r += 1

    for block in panel["blocks"]:
        if block.get("label"):
            ws.cell(r, 1, block["label"]).font = BLOCK_FONT
            r += 1
        cols = block["columns"]
        for j, name in enumerate(cols, start=1):
            cell = ws.cell(r, j, name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        r += 1
        for row in block["rows"]:
            for j, val in enumerate(row, start=1):
                if isinstance(val, (np.floating,)):
                    val = float(val)
                elif isinstance(val, (np.integer,)):
                    val = int(val)
                ws.cell(r, j, val)
            r += 1
        r += 1  # blank line between blocks

    # Thumbnails intentionally disabled (data-only workbook).


# --------------------------------------------------------------------------- #
# panel builders                                                              #
# --------------------------------------------------------------------------- #
def build_fig3d_left():
    from scripts.common import MODEL_CKPTS
    from scripts.plot_localizer_decode import ROIS, ROI_DISPLAY_NAMES, localizer_decode_clustered
    from scripts.get_localizer_decode_ceiling import localizer_decode_ceiling

    ckpts, dropped = _cached_ckpts(MODEL_CKPTS, "decode")
    if dropped:
        print(f"    [skip uncached] {dropped}")
    all_scores = [
        localizer_decode_clustered(ckpt, ROIS, num_splits=1, fwhm_mm=2.0, resolution_mm=1.0)
        for ckpt in ckpts
    ]
    internal_consistency = localizer_decode_ceiling(ROIS, folds=10)
    seeds = [ckpt.split("_sd")[-1].split(".")[0] for ckpt in ckpts]

    columns = ["ROI"] + [f"model sd{s}" for s in seeds] + ["model mean", "human internal-consistency mean", "human internal-consistency SD"]
    rows = []
    for r, roi in enumerate(ROI_DISPLAY_NAMES):
        per_seed = [float(scores[:, r, r].mean()) for scores in all_scores]
        consistency = np.asarray(internal_consistency[r])
        rows.append(
            [roi] + [round(v, 5) for v in per_seed]
            + [round(float(np.mean(per_seed)), 5), round(float(consistency.mean()), 5), round(float(consistency.std()), 5)]
        )
    return dict(
        title="Figure 3d (left) — Functional prediction score per ROI",
        description="Prediction score (Pearson r) decoding each human ROI from the matched model ROI, "
        "averaged over the BOLD Moments and McMahon video datasets. Bars are the mean across 5 "
        "TopoTransform model instances (seeds 42-46, shown individually); horizontal marks are the "
        "mean human internal consistency across 10 split-half estimates; its SD is provided here but is not plotted.",
        blocks=[dict(label=None, columns=columns, rows=rows)],
        thumb=_thumbnail_png(["localizer_decoding_scores_comparison.svg"], "fig3d_left"),
    )


def build_fig3d_right():
    from scripts.plot_localizer_decode import plot_localizer_decode
    from scripts.analysis_utils import CKPT_GROUPS, METHOD_LABELS, resolve_group_names

    # Main-text figure shows only these 4 model types (CLIP/SOM/VideoMAE are supplementary).
    method_order = resolve_group_names(("TopoTransform", "TDANN", "SWAPOPT", "UNOPTIMIZED"))
    per_method = {}
    human_internal_consistency = None
    skipped = []
    for name in method_order:
        ckpts, dropped = _cached_ckpts(CKPT_GROUPS[name], "decode")
        if not ckpts:
            skipped.append(name)
            print(f"    [skip method, no cache] {name}")
            continue
        scores, human = plot_localizer_decode(ckpts, store_dir=None)
        per_method[name] = np.asarray(scores).mean(axis=0)  # mean over ROI -> per seed
        if human_internal_consistency is None:
            human_internal_consistency = np.asarray(human).mean(axis=0)  # per split-half estimate
    method_order = [m for m in method_order if m in per_method]

    max_n = max(len(v) for v in per_method.values())
    columns = ["Model"] + [f"instance {i+1}" for i in range(max_n)] + ["mean"]
    rows = []
    for name in method_order:
        vals = list(per_method[name])
        padded = [round(float(v), 5) for v in vals] + [""] * (max_n - len(vals))
        rows.append([METHOD_LABELS[name]] + padded + [round(float(np.mean(vals)), 5)])

    ci = 1.96 * float(human_internal_consistency.std())
    consistency_rows = [["Human internal-consistency mean", round(float(human_internal_consistency.mean()), 5)],
                        ["Human internal-consistency interval half-width (1.96 SD)", round(ci, 5)]]
    notes = []
    if skipped:
        notes.append("Model type(s) omitted because no cached result is available "
                     f"(not part of the plotted figure): {', '.join(METHOD_LABELS[m] for m in skipped)}.")
    return dict(
        title="Figure 3d (right) — Mean prediction score across models",
        description="Prediction score averaged over the 5 ROIs, per model type. Points are individual "
        "model instances; the shaded band is the human internal-consistency mean +/- 1.96 SD across 10 split-half estimates.",
        notes=notes,
        blocks=[
            dict(label="Per-model scores (mean over ROIs, one column per instance)", columns=columns, rows=rows),
            dict(label="Human internal consistency", columns=["quantity", "value"], rows=consistency_rows),
        ],
        thumb=_thumbnail_png(["localizer_decoding_score_comparison.svg"], "fig3d_right"),
    )


def build_fig3e():
    from scripts.analysis_utils import CKPT_GROUPS, METHOD_LABELS
    from scripts.plot_smoothness import _collect_group_moving, DATASET_NAME, DEFAULT_FWHM

    # Bars as shown in the main-text figure (OneLayer and the SwapOpt/TDANN
    # no-smoothing variants appear only in Supplementary Fig. S4).
    groups = [
        ("UNOPTIMIZED", "no smoothing", 0.0),
        ("UNOPTIMIZED", "smoothing", DEFAULT_FWHM),
        ("SWAPOPT", "smoothing", DEFAULT_FWHM),
        ("TDANN", "smoothing", DEFAULT_FWHM),
        ("TopoTransform", "no smoothing", 0.0),
        ("TopoTransform", "smoothing", DEFAULT_FWHM),
    ]
    human_moving = None
    max_n = 0
    computed = []
    for name, cond, fwhm in groups:
        ckpts = [c for c in CKPT_GROUPS[name] if _cache_exists(_smoothness_cache_name(c, DATASET_NAME, fwhm))]
        if not ckpts:
            print(f"    [skip {name} {cond}, no smoothness cache]")
            continue
        means, human = _collect_group_moving(ckpts, fwhm)
        if human_moving is None:
            human_moving = human
        max_n = max(max_n, len(means))
        computed.append((name, cond, means))

    columns = ["Model", "condition"] + [f"instance {i+1}" for i in range(max_n)] + ["mean", "SD"]
    rows = []
    for name, cond, means in computed:
        padded = [round(float(v), 5) for v in means] + [""] * (max_n - len(means))
        rows.append([METHOD_LABELS[name], cond] + padded + [round(float(np.mean(means)), 5), round(float(np.std(means)), 5)])

    ref = [["Human (moving categories, Moran's I)", round(float(human_moving), 5)]]
    return dict(
        title="Figure 3e — Response smoothness (Moran's I)",
        description=f"Spatial autocorrelation (Moran's I) of response t-values on the model sheet for the "
        f"'{DATASET_NAME}' moving-category localizer, per model type and fMRI-sampling condition. Points are "
        f"individual model instances; the dashed line is the human reference.",
        blocks=[
            dict(label="Model smoothness (one column per instance)", columns=columns, rows=rows),
            dict(label="Human reference", columns=["quantity", "value"], rows=ref),
        ],
        thumb=_thumbnail_png(["smoothness_comparison_bar.svg", "smoothness_comparison_bar_grouped.svg"], "fig3e"),
    )


def _robert_subject_ids():
    ids = []
    for fname in sorted(os.listdir(config.ROBERT_STATS)):
        if fname.endswith(".npy"):
            ids.append(fname.split("_")[0])
    return ids


def _load_robert_tvals_sorted():
    """Per-subject dyn-vs-stat t-maps in a deterministic (sorted-by-filename) order."""
    t_vals, ids = [], []
    for fname in sorted(os.listdir(config.ROBERT_STATS)):
        if fname.endswith(".npy"):
            t_vals.append(np.load(os.path.join(config.ROBERT_STATS, fname)))
            ids.append(fname.split("_")[0])
    return np.array(t_vals), ids


def _motion_roi_props(ckpts, rois):
    """Reproduce plot_localizer_motion.plot_all_rois model/human proportions per ROI."""
    from scripts.get_localizers import get_localizer_model, get_localizer_human
    from scripts.analysis_utils import collect_localizer_tvals

    masks_models = [[] for _ in rois]
    for ckpt in ckpts:
        masks_model = get_localizer_model(rois, ckpt)
        for r in range(len(rois)):
            masks_models[r].append(masks_model[r])
    masks_human = get_localizer_human(rois)
    all_t_vals = collect_localizer_tvals(ckpts, dataset="robert", ret_merged=True, verbose=False)
    t_vals_robert, subj_ids = _load_robert_tvals_sorted()

    model_props, human_props = [], []
    for r in range(len(rois)):
        per_ckpt = []
        for t_vals, mask in zip(all_t_vals, masks_models[r]):
            tv = [t_val[m] for t_val, m in zip(t_vals, mask)][0]
            per_ckpt.append(float((tv > 0).mean()) if tv.size else 0.5)
        per_ckpt = np.nan_to_num(np.array(per_ckpt), nan=0.5)
        per_subj = np.array([float((tv[masks_human[r]] > 0).mean()) for tv in t_vals_robert])
        model_props.append(per_ckpt)
        human_props.append(per_subj)
    return model_props, human_props, subj_ids


def build_fig4c():
    from scripts.common import MODEL_CKPTS

    rois = ["face", "body", "place", "mt", "v6", "psts"]
    display = ["Face", "Body", "Place", "MT", "V6", "pSTS"]
    ckpts, dropped = _cached_ckpts(MODEL_CKPTS, "localizers")
    if dropped:
        print(f"    [skip uncached] {dropped}")
    seeds = [ckpt.split("_sd")[-1].split(".")[0] for ckpt in ckpts]
    model_props, human_props, subj_ids = _motion_roi_props(ckpts, rois)

    model_cols = ["ROI"] + [f"model sd{s}" for s in seeds] + ["model mean", "model SD"]
    model_rows = []
    for r, roi in enumerate(display):
        mp = model_props[r]
        model_rows.append([roi] + [round(float(v), 5) for v in mp] + [round(float(mp.mean()), 5), round(float(mp.std()), 5)])

    human_cols = ["ROI"] + [f"human {sid}" for sid in subj_ids] + ["human mean", "human SD"]
    human_rows = []
    for r, roi in enumerate(display):
        hp = human_props[r]
        human_rows.append([roi] + [round(float(v), 5) for v in hp] + [round(float(hp.mean()), 5), round(float(hp.std()), 5)])

    return dict(
        title="Figure 4c — Proportion of motion-selective units per ROI",
        description="Proportion of motion-selective units/voxels (dynamic > static, t > 0) in each ROI, "
        "human vs. model. Dots are individual data points: model = the 5 TopoTransform instances (seeds 42-46); "
        "human = the individual subjects of Robert et al. 2023.",
        notes=[
            "SAGER note: Human data are reused from Robert et al. 2023 (J Neurosci; object kinematograms), "
            "a sample of n=15 (40% women, i.e. 6 female / 9 male). The source reports only the aggregate sex "
            "composition; per-subject sex is not published and is not available in the released subject-wise "
            "t-statistic maps, so the individual human data points cannot be disaggregated by sex. Model data "
            "points are artificial-network instances (sex not applicable).",
        ],
        blocks=[
            dict(label="Model (one column per instance)", columns=model_cols, rows=model_rows),
            dict(label="Human subjects (one column per subject)", columns=human_cols, rows=human_rows),
        ],
        thumb=_thumbnail_png(
            ["localizer_motion/topotransform/localizer_tvals_comparison.svg",
             "localizer_motion/localizer_tvals_comparison.svg"], "fig4c"),
    )


def build_fig4d():
    from scripts.analysis_utils import CKPT_GROUPS, METHOD_LABELS, resolve_group_names, collect_localizer_tvals
    from scripts.plot_localizer_motion import plot_all_rois

    rois = ["face", "body", "place", "mt", "v6", "psts"]
    # Main-text figure shows only these 4 model types (CLIP/SOM/VideoMAE are supplementary).
    method_order = resolve_group_names(("TopoTransform", "TDANN", "SWAPOPT", "UNOPTIMIZED"))
    per_method = {}
    max_n = 0
    skipped = []
    for name in method_order:
        ckpts, dropped = _cached_ckpts(CKPT_GROUPS[name], "localizers")
        if not ckpts:
            skipped.append(name)
            print(f"    [skip method, no cache] {name}")
            continue
        all_t_vals = collect_localizer_tvals(ckpts, dataset="robert", ret_merged=True, verbose=False)
        mae = np.asarray(plot_all_rois(all_t_vals, ckpts, rois, store_dir=None))
        per_method[name] = mae
        max_n = max(max_n, len(mae))
    method_order = [m for m in method_order if m in per_method]

    columns = ["Model"] + [f"instance {i+1}" for i in range(max_n)] + ["mean"]
    rows = []
    for name in method_order:
        mae = per_method[name]
        padded = [round(float(v), 5) for v in mae] + [""] * (max_n - len(mae))
        rows.append([METHOD_LABELS[name]] + padded + [round(float(mae.mean()), 5)])

    notes = []
    if skipped:
        notes.append("Model type(s) omitted because no cached result is available "
                     f"(not part of the plotted figure): {', '.join(METHOD_LABELS[m] for m in skipped)}.")
    return dict(
        title="Figure 4d — Motion-selectivity error vs. humans",
        description="Mean absolute error between model and human ROI-level proportions of motion-selective "
        "units, averaged across ROIs, per model type. Points are individual model instances.",
        notes=notes,
        blocks=[dict(label=None, columns=columns, rows=rows)],
        thumb=_thumbnail_png(["localizer_motion/localizer_motion_mae_comparison.svg"], "fig4d"),
    )


def build_fig4e():
    from scripts.common import MODEL_CKPTS
    from scripts.get_localizers import localizers
    from validate.rois.nsd import get_region_voxels

    ckpts, dropped = _cached_ckpts(MODEL_CKPTS, "localizers")
    if dropped:
        print(f"    [skip uncached] {dropped}")
    t_vals_model = []
    for ckpt in ckpts:
        t_dicts, _, _ = localizers(ckpt, ret_merged=True)
        t_vals_model.append(np.asarray(t_dicts["robert"][0]).flatten())
    t_vals_model = np.array(t_vals_model).mean(axis=0)

    t_vals_human, _ = _load_robert_tvals_sorted()
    t_vals_human = t_vals_human.mean(0)
    high = get_region_voxels(["high-dorsal", "high-ventral", "high-lateral"])
    t_vals_human = t_vals_human[high].flatten()

    def _hist_block(values, label):
        counts, edges = np.histogram(values, bins=50)
        prob = counts / counts.sum()
        rows = [[round(float(edges[i]), 5), round(float(edges[i + 1]), 5),
                 round(float((edges[i] + edges[i + 1]) / 2), 5), int(counts[i]), round(float(prob[i]), 6)]
                for i in range(len(counts))]
        return dict(label=label, columns=["bin left", "bin right", "bin center", "count", "probability"], rows=rows)

    stats = [["Model t-values: mean", round(float(t_vals_model.mean()), 5), "SD", round(float(t_vals_model.std()), 5),
              "n units", int(t_vals_model.size)],
             ["Human t-values: mean", round(float(t_vals_human.mean()), 5), "SD", round(float(t_vals_human.std()), 5),
              "n vertices", int(t_vals_human.size)]]

    return dict(
        title="Figure 4e — Distribution of motion-selectivity t-values",
        description="Histogram (probability, 50 bins) of dynamic-vs-static t-values across the whole model "
        "cortical sheet (main panel) and across human high-level visual cortex (inset). Averaged over the 5 "
        "TopoTransform instances (model) and over subjects (human, Robert et al. 2023).",
        blocks=[
            dict(label="Summary", columns=["", "", "", "", ""], rows=stats),
            _hist_block(t_vals_model, "Model t-value histogram"),
            _hist_block(t_vals_human, "Human t-value histogram"),
        ],
        thumb=_thumbnail_png(["robert_tval_distribution.svg"], "fig4e"),
    )


def build_fig2e():
    import glob
    from scripts.plot_alignment_task_loss_timeseries import DEFAULT_PATTERN, compute_timeseries

    ckpt_paths = sorted(glob.glob(DEFAULT_PATTERN))
    if not ckpt_paths:
        raise FileNotFoundError(f"No per-epoch checkpoints for pattern: {DEFAULT_PATTERN}")

    def _epoch_cached(name):
        checks = [
            f"neural_alignment_splits1_{name}",
            f"task_performance_imagenet_{name}",
            f"task_performance_ssv2_{name}",
            f"topographic_loss_{name}_bs32_spb2048_maxb8_devicecpu",
        ]
        return all(_cache_exists(c) for c in checks)

    kept = [p for p in ckpt_paths if _epoch_cached(Path(p).name)]
    dropped = [Path(p).name for p in ckpt_paths if p not in kept]
    if dropped:
        print(f"    [skip uncached epochs] {dropped}")
    ckpt_paths = kept
    if not ckpt_paths:
        raise RuntimeError("No epoch checkpoints have a complete cache set.")
    epochs, neural, task, topo = compute_timeseries(
        ckpt_paths, num_splits=1, batch_size=32, samples_per_batch=2048, max_batches=8, device="cpu"
    )
    columns = ["epoch", "topographic loss", "neural alignment (normalized by internal consistency)", "task performance (avg)"]
    rows = [[int(e), round(float(l), 6), round(float(n), 6), round(float(t), 6)]
            for e, l, n, t in zip(epochs, topo, neural, task)]
    return dict(
        title="Figure 2e — Task performance and neural alignment over training",
        description="Across training epochs of the invertible transform: topographic loss, system-level neural "
        "alignment normalized by internal consistency, and task performance averaged over object (ImageNet) and "
        "action (SSv2) recognition.",
        blocks=[dict(label=None, columns=columns, rows=rows)],
        thumb=_thumbnail_png(["alignment_task_loss_timeseries.svg"], "fig2e"),
    )


def _binned_similarity_stats(features_unit_by_stim, positions, n_bins=11, max_distance=65, subsample=10000, seed=42):
    """Per-bin response-correlation statistics vs cortical distance (Fig. 2c)."""
    rng = np.random.default_rng(seed)
    feats = features_unit_by_stim
    pos = np.asarray(positions)
    if subsample is not None and feats.shape[0] > subsample:
        idx = rng.choice(feats.shape[0], subsample, replace=False)
        feats = feats[idx]
        pos = pos[idx]
    feats = (feats - feats.mean(axis=1, keepdims=True)) / (feats.std(axis=1, keepdims=True) + 1e-10)
    dist = ((pos[:, None, :] - pos[None, :, :]) ** 2).sum(-1) ** 0.5
    sim = (feats @ feats.T) / feats.shape[1]
    tri = np.triu_indices_from(dist, k=1)
    d, s = dist[tri], sim[tri]
    keep = d <= max_distance
    d, s = d[keep], s[keep]
    bins = np.linspace(d.min(), d.max(), n_bins + 1)
    centers = (bins[:-1] + bins[1:]) / 2
    which = np.digitize(d, bins)
    rows = []
    for i in range(1, len(bins)):
        vals = s[which == i]
        if vals.size == 0:
            rows.append([round(float(centers[i - 1]), 3), "", "", "", "", "", 0])
        else:
            rows.append([
                round(float(centers[i - 1]), 3), round(float(np.mean(vals)), 5), round(float(np.std(vals)), 5),
                round(float(np.percentile(vals, 25)), 5), round(float(np.median(vals)), 5),
                round(float(np.percentile(vals, 75)), 5), int(vals.size),
            ])
    return rows


def build_fig2c():
    from scripts.common import MODEL_CKPT
    from scripts.get_validate_features import validate_features

    all_features, positions = validate_features(MODEL_CKPT)
    feats = all_features[0]
    if feats.ndim == 5:
        B, T, C, H, W = feats.shape
        feats = feats.reshape(B * T, C * H * W)
    else:
        B, C, H, W = feats.shape
        feats = feats.reshape(B, C * H * W)
    feats = feats.numpy().T  # (units, stimuli)
    pos = positions[0].coordinates
    pos = pos.numpy() if hasattr(pos, "numpy") else np.asarray(pos)
    rows = _binned_similarity_stats(feats, pos)
    return dict(
        title="Figure 2c — Model unit response similarity across distance",
        description="Pairwise response correlation between units on the optimized model sheet, binned by "
        "cortical distance (TopoTransform, seed 45). Box plots in the figure summarize each bin; statistics "
        "below are computed on a seeded 10,000-unit subsample.",
        notes=["The published scatter uses an unseeded 10k-unit subsample; per-bin statistics here use a fixed "
               "seed (42) and match the plotted distribution within sampling noise."],
        blocks=[dict(label=None,
                     columns=["distance bin center (mm)", "mean", "SD", "Q1", "median", "Q3", "n pairs"],
                     rows=rows)],
        thumb=_thumbnail_png(["plot_wiring_cost.svg"], "fig2c_left"),
    )


def build_fig2d():
    from scripts.common import MODEL_CKPT
    from scripts.plot_wiring_cost_fmri import (
        _collect_fmri_batches, _prepare_human_fmri, _prepare_model_fmri_features,
        _compute_binned_similarity, _cached_geodesic_distances_high, _normalize_features, _flatten_features,
    )
    from validate.smoothness import NSD_HIGH

    import pickle

    seed, subsample, n_bins, max_distance = 42, 2000, 11, 65
    raw_path = OUT_DIR / "_fig2c_right_raw.pkl"
    if raw_path.exists():
        with open(raw_path, "rb") as fh:
            series = pickle.load(fh)
        print("    [fig2c_right] loaded cached inference result")
    else:
        np.random.seed(seed)
        stim_batches, fmri_batches = _collect_fmri_batches(max_batches=None, seed=seed)
        human = _prepare_human_fmri(fmri_batches)
        model_feats, model_pos = _prepare_model_fmri_features(stim_batches, MODEL_CKPT, do_transform=True)
        pre_feats, pre_pos = _prepare_model_fmri_features(stim_batches, MODEL_CKPT, do_transform=False)

        def _series(flat, pos, geodesic_idx=None):
            flat = _flatten_features(flat)
            n = flat.shape[1]
            idx = np.arange(n)
            if subsample is not None and n > subsample:
                idx = np.random.choice(n, subsample, replace=False)
            flat = flat[:, idx]
            norm = _normalize_features(flat.T)
            if geodesic_idx is not None:
                glob_idx = np.where(NSD_HIGH)[0][idx]
                dist = _cached_geodesic_distances_high(glob_idx, seed=seed)
            else:
                p = pos[idx]
                dist = np.linalg.norm(p[:, None, :] - p[None, :, :], axis=-1)
            return _compute_binned_similarity(norm, dist, n_bins=n_bins, max_distance=max_distance)

        series = {
            "model": _series(model_feats, model_pos),
            "pre": _series(pre_feats, pre_pos),
            "human": _series(human, None, geodesic_idx=True),
        }
        with open(raw_path, "wb") as fh:
            pickle.dump(series, fh)

    m_c, m_m, m_s, m_n = series["model"]
    p_c, p_m, p_s, p_n = series["pre"]
    h_c, h_m, h_s, h_n = series["human"]

    def _rows(c, m, s, n):
        return [[round(float(c[i]), 3), round(float(m[i]), 5), round(float(s[i]), 5), int(n[i])]
                for i in range(len(c))]

    cols = ["distance bin center (mm)", "mean correlation", "SD", "n pairs"]
    return dict(
        title="Figure 2d — fMRI response similarity across distance",
        description="Response correlation as a function of cortical distance for simulated model fMRI (post- and "
        "pre-transform) and human fMRI (BOLD Moments; geodesic distance on fsaverage5). Lines are per-bin mean; "
        "shaded regions are +/- 1 SD.",
        blocks=[
            dict(label="Model fMRI (post-transform)", columns=cols, rows=_rows(m_c, m_m, m_s, m_n)),
            dict(label="Initial model fMRI (pre-transform)", columns=cols, rows=_rows(p_c, p_m, p_s, p_n)),
            dict(label="Human fMRI", columns=cols, rows=_rows(h_c, h_m, h_s, h_n)),
        ],
        thumb=_thumbnail_png(["plot_wiring_cost_fmri.svg"], "fig2c_right"),
    )


FIG2B_SEED = 35  # chosen Fig 2b seed; set to the seed you pick from `python -m scripts.plot_fig2b`


def build_fig2b():
    """Exemplary unit-activation traces (Fig 2b right), from the shared seeded selector."""
    from scripts.common import MODEL_CKPT
    from scripts.get_validate_features import validate_features
    from scripts.plot_fig2b import fig2b_unit_data

    import json

    # Digitized traces live in the distributed cache; source_data/ is a local override.
    candidates = [
        OUT_DIR / "fig2b_traces.json",
        DEBUG_DIR / "fig2b_traces.json",
        PLOTS_DIR / "fig2b" / "fig2b_traces.json",
    ]
    raw = next((p for p in candidates if p.exists()), None)
    if raw is None:
        raise FileNotFoundError(
            "Missing Fig 2b traces (fig2b_traces.json); looked in: "
            + ", ".join(str(p) for p in candidates)
        )
    print(f"    [fig2b] traces from {raw}")
    payload = json.load(open(raw))
    rows_in = payload["rows"] if isinstance(payload, dict) else payload
    labels = ["Unit 1 (probe)", "Unit 2 (correlated)", "Unit 3 (decorrelated)"]
    columns = ["clip", "frame", "time index"] + labels
    if isinstance(payload, dict):  # rows = [clip, frame, time_index, u1, u2, u3]
        rows = [[r[0], int(r[1]), int(r[2]),
                 round(float(r[3]), 6), round(float(r[4]), 6), round(float(r[5]), 6)] for r in rows_in]
    else:  # legacy rows = [time_index, clip, frame, u1, u2, u3]
        rows = [[r[1], int(r[2]), int(r[0]),
                 round(float(r[3]), 6), round(float(r[4]), 6), round(float(r[5]), 6)] for r in rows_in]

    return dict(
        title="Figure 2b (right) — Exemplary unit activations over stimuli",
        description="Activations of three example units on the model cortical sheet (two spatially proximal, "
        "one distant) across five video clips (S1-S5), 12 frames per clip.",
        notes=["Values digitized directly from the published figure: the three trace polylines were read at full "
               "coordinate precision from the vector paths in figs/F2.pdf and mapped to data units using the "
               "panel's y-axis ticks, so these numbers correspond exactly to the printed panel."],
        blocks=[dict(label="Activation traces", columns=columns, rows=rows)],
        thumb=None,
    )


## ------------------------------------------------------------------------- ##
## Supplementary figure panels                                               ##
## ------------------------------------------------------------------------- ##
_SI_DECODE_ORDER = ("TopoTransform", "TDANN", "SWAPOPT", "UNOPTIMIZED", "ONELAYER")


def _decode_per_method(order_names):
    from scripts.plot_localizer_decode import plot_localizer_decode
    from scripts.analysis_utils import CKPT_GROUPS, METHOD_LABELS, resolve_group_names

    order = resolve_group_names(order_names)
    per, human, skipped = {}, None, []
    for name in order:
        ckpts, _ = _cached_ckpts(CKPT_GROUPS[name], "decode")
        if not ckpts:
            skipped.append(name)
            print(f"    [skip method, no cache] {name}")
            continue
        scores, hu = plot_localizer_decode(ckpts, store_dir=None)
        per[name] = np.asarray(scores).mean(axis=0)
        if human is None:
            human = np.asarray(hu).mean(axis=0)
    order = [m for m in order if m in per]
    return order, per, human, skipped, METHOD_LABELS


def build_si_S4a():
    order, per, human, skipped, LABELS = _decode_per_method(_SI_DECODE_ORDER)
    max_n = max(len(v) for v in per.values())
    cols = ["Model"] + [f"instance {i+1}" for i in range(max_n)] + ["mean"]
    rows = []
    for m in order:
        vals = [round(float(v), 5) for v in per[m]]
        rows.append([LABELS[m]] + vals + [""] * (max_n - len(vals)) + [round(float(np.mean(per[m])), 5)])
    ci = 1.96 * float(human.std())
    return dict(
        title="Figure S4a — ROI-wise functional alignment across models (incl. OneLayer)",
        description="ROI-averaged localizer decoding score (Pearson R) per model type, extended set including "
        "OneLayer (cf. Fig. 3d). Points are individual model instances; the band is the human internal-consistency mean +/- 1.96 SD across 10 split-half estimates.",
        blocks=[
            dict(label="Per-model scores (one column per instance)", columns=cols, rows=rows),
            dict(label="Human internal consistency", columns=["quantity", "value"],
                 rows=[["mean", round(float(human.mean()), 5)], ["95% CI half-width", round(ci, 5)]]),
        ],
        thumb=None,
    )


def build_si_S4b():
    from scripts.analysis_utils import CKPT_GROUPS, METHOD_LABELS
    from scripts.plot_smoothness import _collect_group_moving, DATASET_NAME, DEFAULT_FWHM

    groups = [
        ("UNOPTIMIZED", "no smoothing", 0.0), ("UNOPTIMIZED", "smoothing", DEFAULT_FWHM),
        ("SWAPOPT", "no smoothing", 0.0), ("SWAPOPT", "smoothing", DEFAULT_FWHM),
        ("TDANN", "no smoothing", 0.0), ("TDANN", "smoothing", DEFAULT_FWHM),
        ("TopoTransform", "no smoothing", 0.0), ("TopoTransform", "smoothing", DEFAULT_FWHM),
        ("ONELAYER", "no smoothing", 0.0), ("ONELAYER", "smoothing", DEFAULT_FWHM),
    ]
    human_moving, max_n, computed = None, 0, []
    for name, cond, fwhm in groups:
        ckpts = [c for c in CKPT_GROUPS[name] if _cache_exists(_smoothness_cache_name(c, DATASET_NAME, fwhm))]
        if not ckpts:
            continue
        means, human = _collect_group_moving(ckpts, fwhm)
        if human_moving is None:
            human_moving = human
        max_n = max(max_n, len(means))
        computed.append((name, cond, means))
    cols = ["Model", "condition"] + [f"instance {i+1}" for i in range(max_n)] + ["mean", "SD"]
    rows = []
    for name, cond, means in computed:
        vals = [round(float(v), 5) for v in means]
        rows.append([METHOD_LABELS[name], cond] + vals + [""] * (max_n - len(vals))
                    + [round(float(np.mean(means)), 5), round(float(np.std(means)), 5)])
    return dict(
        title="Figure S4b — Response smoothness, grouped no-smoothing vs smoothing (incl. OneLayer)",
        description="Spatial autocorrelation (Moran's I) for the pitcher moving-category localizer, all five "
        "model types with both fMRI-sampling conditions (cf. Fig. 3e).",
        blocks=[
            dict(label="Model smoothness (one column per instance)", columns=cols, rows=rows),
            dict(label="Human reference", columns=["quantity", "value"],
                 rows=[["Human (moving categories, Moran's I)", round(float(human_moving), 5)]]),
        ],
        thumb=_thumbnail_png(["smoothness_comparison_bar_grouped.svg"], "siS4b"),
    )


def build_si_S4c():
    from scripts.analysis_utils import CKPT_GROUPS, METHOD_LABELS, resolve_group_names, collect_localizer_tvals
    from scripts.plot_localizer_motion import plot_all_rois

    rois = ["face", "body", "place", "mt", "v6", "psts"]
    order = resolve_group_names(_SI_DECODE_ORDER)
    per, max_n = {}, 0
    for name in order:
        ckpts, _ = _cached_ckpts(CKPT_GROUPS[name], "localizers")
        if not ckpts:
            print(f"    [skip method, no cache] {name}")
            continue
        all_t = collect_localizer_tvals(ckpts, dataset="robert", ret_merged=True, verbose=False)
        mae = np.asarray(plot_all_rois(all_t, ckpts, rois, store_dir=None))
        per[name] = mae
        max_n = max(max_n, len(mae))
    order = [m for m in order if m in per]
    cols = ["Model"] + [f"instance {i+1}" for i in range(max_n)] + ["mean"]
    rows = []
    for m in order:
        vals = [round(float(v), 5) for v in per[m]]
        rows.append([METHOD_LABELS[m]] + vals + [""] * (max_n - len(vals)) + [round(float(np.mean(per[m])), 5)])
    return dict(
        title="Figure S4c — Motion-selectivity error across models (incl. OneLayer)",
        description="ROI-averaged mean absolute error of motion-selective proportion vs. humans, extended set "
        "including OneLayer (cf. Fig. 4d). Points are individual model instances.",
        blocks=[dict(label=None, columns=cols, rows=rows)],
        thumb=None,
    )


def build_si_S9():
    import glob
    from scripts.plot_alignment_task_loss_timeseries2 import DEFAULT_PATTERN, compute_timeseries

    def _epoch_cached(name):
        return all(_cache_exists(c) for c in [
            f"neural_alignment_splits1_{name}", f"task_performance_imagenet_{name}",
            f"task_performance_ssv2_{name}", f"topographic_loss_{name}_bs32_spb2048_maxb8_devicecpu"])

    paths = [p for p in sorted(glob.glob(DEFAULT_PATTERN)) if _epoch_cached(Path(p).name)]
    if not paths:
        raise RuntimeError("No epoch checkpoints have a complete cache set for S9.")
    epochs, neural, task, topo, pre_neural, pre_task = compute_timeseries(
        paths, num_splits=1, batch_size=32, samples_per_batch=2048, max_batches=8, device="cpu")
    rows = [[int(e), round(float(n), 6), round(float(t), 6), round(float(l), 6)]
            for e, n, t, l in zip(epochs, neural, task, topo)]
    return dict(
        title="Figure S9 — Neural alignment & task performance over training, pre vs post transform",
        description="Post-transform neural alignment (fraction of ceiling) and task performance across epochs, "
        "with the pre-transform (untransformed) baselines (cf. Fig. 2d).",
        blocks=[
            dict(label="Post-transform, per epoch",
                 columns=["epoch", "neural alignment (post)", "task performance (post)", "topographic loss"], rows=rows),
            dict(label="Pre-transform baseline (untransformed model)",
                 columns=["quantity", "value"],
                 rows=[["neural alignment (pre)", round(float(pre_neural), 6)],
                       ["task performance (pre)", round(float(pre_task), 6)]]),
        ],
        thumb=None,
    )


def build_si_S10():
    import pickle
    store = Path(
        os.environ.get("HIERARCHY_ALIGNMENT_CACHE", DEBUG_DIR / "hierarchy_alignment")
    )
    from scripts.common import MODEL_CKPT
    tag = MODEL_CKPT.replace(".", "_")
    cache_file = store / f"hierarchy_alignment_ckpt{tag}_N6_splits1_modeposttransform_v4.p"
    if not cache_file.exists():
        raise FileNotFoundError(f"No cached hierarchy alignment: {cache_file}")
    with open(cache_file, "rb") as fh:
        res = pickle.load(fh)
    from scipy.stats import pearsonr
    dist = np.asarray(res["aligned_distance"])
    geo = np.asarray(res["geodesic"])
    score = np.asarray(res["aligned_score"])
    keep = (dist < 1000.0) & (score > 0.5)
    dist, geo, score = dist[keep], geo[keep], score[keep]
    r, p = pearsonr(dist, geo)
    rows = [[round(float(dist[i]), 4), round(float(geo[i]), 4), round(float(score[i]), 5)] for i in range(len(dist))]
    return dict(
        title="Figure S10 — Posterior–anterior model↔human correspondence",
        description="Each point is a matched model-sheet segment / human high-level-cortex voxel pair with "
        "decoding score > 0.5 (cf. Supplementary Fig.; n_segments=6, seed-45 model).",
        notes=[f"Pearson r = {r:.3f}, p = {p:.3g} across {len(dist)} points."],
        blocks=[dict(label=None,
                     columns=["model sheet segment distance (mm)", "human cortex geodesic distance (mm)", "decoding score (r)"],
                     rows=rows)],
        thumb=None,
    )


def _motion_props_scaled(ckpts, rois, scale):
    """Model motion-selective proportion per ROI, with ROI-defining t-thresholds scaled by `scale`.

    Re-thresholds the cached localizer t/p values (p<0.001 kept fixed; only the t-threshold is
    scaled, matching the manuscript's 125%/75% conditions). Cache-only, no recompute.
    """
    from scripts.get_localizers import localizers
    from scripts.localizer_registry import get_localizer_result_key, get_roi_t_threshold
    from scripts.common import LOCALIZER_T_THRESHOLD, LOCALIZER_P_THRESHOLD

    per = [[] for _ in rois]
    for ckpt in ckpts:
        tvd, pvd, _ = localizers(ckpt, ret_merged=True)
        robert = np.asarray(tvd["robert"][0]).flatten()
        for i, roi in enumerate(rois):
            key = get_localizer_result_key(roi)
            t = np.asarray(tvd[key][0]).flatten()
            p = np.asarray(pvd[key][0]).flatten()
            base_t = get_roi_t_threshold(roi, LOCALIZER_T_THRESHOLD)
            mask = (p < LOCALIZER_P_THRESHOLD) & (t > base_t * scale)
            per[i].append(float((robert[mask] > 0).mean()) if int(mask.sum()) else 0.5)
    return [np.nan_to_num(np.array(v), nan=0.5) for v in per]


def _build_S8(scale, letter, pct):
    from scripts.common import MODEL_CKPTS

    rois = ["face", "body", "place", "mt", "v6", "psts"]
    display = ["Face", "Body", "Place", "MT", "V6", "pSTS"]
    ckpts, _ = _cached_ckpts(MODEL_CKPTS, "localizers")
    seeds = [c.split("_sd")[-1].split(".")[0] for c in ckpts]
    model = _motion_props_scaled(ckpts, rois, scale)
    _, human_props, subj_ids = _motion_roi_props(ckpts, rois)  # human is threshold-independent

    m_cols = ["ROI"] + [f"model sd{s}" for s in seeds] + ["model mean", "model SD"]
    m_rows = [[display[r]] + [round(float(v), 5) for v in model[r]]
              + [round(float(model[r].mean()), 5), round(float(model[r].std()), 5)] for r in range(len(rois))]
    h_cols = ["ROI"] + [f"human {s}" for s in subj_ids] + ["human mean", "human SD"]
    h_rows = [[display[r]] + [round(float(v), 5) for v in human_props[r]]
              + [round(float(human_props[r].mean()), 5), round(float(human_props[r].std()), 5)] for r in range(len(rois))]
    return dict(
        title=f"Figure S8{letter} — Motion-selective proportion at {pct} t-threshold",
        description=f"Proportion of motion-selective units per ROI with the model ROI-defining t-value threshold "
        f"scaled to {pct} of the main-text value (cf. Fig. 4c). Human values are threshold-independent "
        f"(fixed atlas ROIs), identical to Fig. 4c.",
        blocks=[
            dict(label="Model (one column per instance)", columns=m_cols, rows=m_rows),
            dict(label="Human subjects (one column per subject)", columns=h_cols, rows=h_rows),
        ],
        thumb=None,
    )


def build_si_S8a():
    return _build_S8(1.25, "a", "125%")


def build_si_S8b():
    return _build_S8(0.75, "b", "75%")


## ------------------------------------------------------------------------- ##
## Supplementary rebuttal panels: S2 (CLIP), S3 (other models), S6 (motion)   ##
## ------------------------------------------------------------------------- ##
def _decode_cache_name_rois(ckpt, rois):
    from scripts.plot_localizer_decode import _CLUSTER_K

    rc = hashlib.md5("_".join(sorted(rois)).encode()).hexdigest()[:8]
    kc = hashlib.md5("_".join(f"{k}:{v}" for k, v in sorted(_CLUSTER_K.items())).encode()).hexdigest()[:8]
    return f"localizer_decode_clustered_splits1_{ckpt}_rois{rc}_k{kc}_fwhm2.0_res1.0"


def _moving_smoothness(ckpt):
    from scripts.get_smoothness import smoothness

    res = smoothness(ckpt, "pitcher", fwhm_mm=2.0, resolution_mm=1.0)
    moving = [k for k in res if "moving" in k.lower()] or list(res)
    model = float(np.mean([res[k]["model_smoothness"] for k in moving]))
    human = float(np.mean([res[k]["human_smoothness"] for k in moving]))
    return model, human


def _norm_decode_per_roi(group, ckpts):
    """Per-ROI decoding R normalized by human internal consistency, seed-averaged."""
    from scripts.plot_localizer_decode import localizer_decode_clustered
    from scripts.get_localizer_decode_ceiling import localizer_decode_ceiling
    from scripts.plot_rebuttal_experiments import (
        _decode_rois_for_group, _decode_run_rois_for_group, REBUTTAL_DECODE_ROIS)

    decode_rois = _decode_rois_for_group(group)
    run_rois = _decode_run_rois_for_group(group)
    ceil = localizer_decode_ceiling(REBUTTAL_DECODE_ROIS, folds=10)
    cmap = {roi: float(np.asarray(ceil[REBUTTAL_DECODE_ROIS.index(roi)]).mean()) for roi in REBUTTAL_DECODE_ROIS}

    usable = [c for c in ckpts if _cache_exists(_decode_cache_name_rois(c, run_rois))]
    per_roi = {}
    for roi in decode_rois:
        vals = []
        for ckpt in usable:
            sc = localizer_decode_clustered(ckpt, run_rois, num_splits=1, fwhm_mm=2.0, resolution_mm=1.0)
            R = float(sc[:, run_rois.index(roi), run_rois.index(roi)].mean())
            vals.append(R / cmap[roi])
        per_roi[roi] = float(np.mean(vals)) if vals else None
    return decode_rois, per_roi


def _build_rebuttal_panel(title, description, group_order, add_models):
    from scripts.plot_rebuttal_experiments import _existing_rebuttal_groups, REBUTTAL_METHOD_LABELS

    groups = _existing_rebuttal_groups(add_models=add_models)
    all_rois = ["face", "place", "body", "v6", "psts"]
    disp = ["face", "place", "body", "V6", "pSTS"]

    dec_rows, human_ref = [], None
    for g in group_order:
        if g not in groups:
            continue
        drois, pernorm = _norm_decode_per_roi(g, groups[g])
        row = [REBUTTAL_METHOD_LABELS[g]] + [
            round(pernorm[r], 6) if pernorm.get(r) is not None else "not applied"
            for r in all_rois
        ]
        present = [pernorm[r] for r in drois if pernorm.get(r) is not None]
        row.append(round(float(np.mean(present)), 6) if present else "")
        row.append(
            round(float(np.std(present, ddof=1) / np.sqrt(len(present))), 6)
            if len(present) > 1 else "not applied"
        )
        dec_rows.append(row)
    dec_cols = ["Model"] + disp + ["mean (normalized R)", "SEM across plotted ROI points"]

    sm_rows, max_n, sm = [], 0, {}
    for g in group_order:
        if g not in groups:
            continue
        vals = []
        for c in groups[g]:
            m, h = _moving_smoothness(c)
            vals.append(m)
            if human_ref is None:
                human_ref = h
        sm[g] = vals
        max_n = max(max_n, len(vals))
    for g in group_order:
        if g not in sm:
            continue
        vals = sm[g]
        sem = (
            round(float(np.std(vals, ddof=1) / np.sqrt(len(vals))), 6)
            if len(vals) > 1 else "not applied"
        )
        sm_rows.append([REBUTTAL_METHOD_LABELS[g]] + [round(v, 6) for v in vals]
                       + ["not applied"] * (max_n - len(vals))
                       + [round(float(np.mean(vals)), 6), sem])
    sm_cols = ["Model"] + [f"instance {i+1}" for i in range(max_n)] + ["mean", "SEM across instances"]

    return dict(
        title=title,
        description=(
            description
            + " The localizer-decoding R values are normalized by human internal consistency. "
              "Bar heights are means over the displayed points; plotted error bars are SEM."
        ),
        blocks=[
            dict(label="Localizer decoding — normalized R (per-ROI R / human internal consistency); the per-ROI values are the plotted points",
                 columns=dec_cols, rows=dec_rows),
            dict(label="Spatial smoothness — Moran's I (moving categories, with fMRI sampling); one column per instance",
                 columns=sm_cols, rows=sm_rows),
            dict(label="Human reference", columns=["quantity", "value"],
                 rows=[["Human smoothness (moving, Moran's I)", round(float(human_ref), 6)]]),
        ],
        thumb=None,
    )


def build_si_S2():
    return _build_rebuttal_panel(
        "Figure S2 — TopoTransform applied to CLIP",
        "Localizer decoding (normalized R) and spatial smoothness (Moran's I) for VJEPA and CLIP, each "
        "transformed vs. untransformed. Decoding points are per ROI (5 ROIs for VJEPA, 3 for CLIP).",
        ["UNOPTIMIZED", "TopoTransform", "CLIP_RAW", "CLIP"], None,
    )


def build_si_S3():
    return _build_rebuttal_panel(
        "Figure S3 — Comparison to static topographic vision models",
        "Localizer decoding (normalized R) and spatial smoothness for VJEPA-TopoTransform, CLIP-TopoTransform, "
        "TDANN, LLCNN, TopoNets, and VJEPA-SwapOpt.",
        ["TopoTransform", "CLIP", "TDANN", "LLCNN", "TOPONETS", "SWAPOPT"], "all",
    )


def build_si_S6():
    from scripts.analysis_utils import CKPT_GROUPS, resolve_group_names, METHOD_LABELS

    rois = ["face", "body", "place", "mt", "v6", "psts"]
    display = ["Face", "Body", "Place", "MT", "V6", "pSTS"]
    order = resolve_group_names(("TopoTransform", "UNOPTIMIZED", "SWAPOPT", "TDANN"))
    human_shared, subj_ids, blocks = None, None, []
    for name in order:
        ckpts, _ = _cached_ckpts(CKPT_GROUPS[name], "localizers")
        if not ckpts:
            print(f"    [skip method, no cache] {name}")
            continue
        model_props, human_props, sids = _motion_roi_props(ckpts, rois)
        if human_shared is None:
            human_shared, subj_ids = human_props, sids
        seeds = [c.split("_sd")[-1].split(".")[0] for c in ckpts]
        cols = ["ROI"] + [f"sd{s}" for s in seeds] + ["mean", "SD"]
        rows = [[display[r]] + [round(float(v), 6) for v in model_props[r]]
                + [round(float(model_props[r].mean()), 6), round(float(model_props[r].std()), 6)] for r in range(len(rois))]
        blocks.append(dict(label=f"{METHOD_LABELS[name]} — model motion-selective proportion per ROI (one column per instance)",
                           columns=cols, rows=rows))
    h_cols = ["ROI"] + [f"human {s}" for s in subj_ids] + ["mean", "SD"]
    h_rows = [[display[r]] + [round(float(v), 6) for v in human_shared[r]]
              + [round(float(human_shared[r].mean()), 6), round(float(human_shared[r].std()), 6)] for r in range(len(rois))]
    blocks.append(dict(label="Human motion-selective proportion per ROI (shared across models, one column per subject)",
                       columns=h_cols, rows=h_rows))
    return dict(
        title="Figure S6 — Motion-selective proportion per ROI across models",
        description="Proportion of motion-selective units per ROI (model vs. human) for four model types "
        "(cf. Fig. 4c). Human values are shared across the model panels.",
        blocks=blocks,
        thumb=None,
    )


CHEAP = [
    ("Fig 2e", build_fig2e),
    ("Fig 3d left", build_fig3d_left),
    ("Fig 3d right", build_fig3d_right),
    ("Fig 3e", build_fig3e),
    ("Fig 4c", build_fig4c),
    ("Fig 4d", build_fig4d),
    ("Fig 4e", build_fig4e),
]
def build_si_S1():
    """S1: VJEPA block -> monkey V1/V4/IT alignment (THINGS Ventral Spiking Dataset)."""
    import json
    candidates = [OUT_DIR / "figS1_curves.json", DEBUG_DIR / "figS1_curves.json"]
    raw = next((p for p in candidates if p.exists()), None)
    if raw is None:
        raise FileNotFoundError("Missing figS1_curves.json; looked in: " + ", ".join(map(str, candidates)))
    print(f"    [figS1] curves from {raw}")
    art = json.load(open(raw))
    blocks, curves = art["blocks"], art["curves"]
    regions = ["V1", "V4", "IT"]

    plotted_rows = [[int(b)] + [round(float(curves[r][i]), 6) for r in regions]
                    for i, b in enumerate(blocks)]

    return dict(
        title="Figure S1 — VJEPA block alignment to monkey V1/V4/IT",
        description="Variance explained normalized by internal consistency when linearly regressing each VJEPA block onto monkey "
        "electrode responses (THINGS Ventral Spiking Dataset; V1/V4/IT, two macaques). Blocks 14-23 are "
        "highlighted in the figure as the range used for the cortical sheet.",
        blocks=[dict(
            label="Normalized variance explained (per model block)",
            columns=["model block"] + regions,
            rows=plotted_rows,
        )],
        thumb=None,
    )


SI = [
    ("Fig S1", build_si_S1),
    ("Fig S2", build_si_S2),
    ("Fig S3", build_si_S3),
    ("Fig S4a", build_si_S4a),
    ("Fig S4b", build_si_S4b),
    ("Fig S4c", build_si_S4c),
    ("Fig S6", build_si_S6),
    ("Fig S8a", build_si_S8a),
    ("Fig S8b", build_si_S8b),
    ("Fig S9", build_si_S9),
    ("Fig S10", build_si_S10),
]
HEAVY = [
    ("Fig 2b", build_fig2b),
    ("Fig 2c", build_fig2c),
    ("Fig 2d", build_fig2d),
]
# Desired left-to-right tab order in the final workbook.
ORDER = ["Fig 2b", "Fig 2c", "Fig 2d", "Fig 2e",
         "Fig 3d left", "Fig 3d right", "Fig 3e", "Fig 4c", "Fig 4d", "Fig 4e",
         "Fig S1", "Fig S2", "Fig S3", "Fig S4a", "Fig S4b", "Fig S4c", "Fig S6",
         "Fig S8a", "Fig S8b", "Fig S9", "Fig S10"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--heavy", action="store_true", help="Also build Fig 2b--d (feature load + inference).")
    ap.add_argument("--only", nargs="*", default=None, help="Build only these tab names.")
    args = ap.parse_args()

    builders = dict(CHEAP + SI + (HEAVY if args.heavy else []))
    if args.only:
        builders = {k: v for k, v in builders.items() if k in set(args.only)}

    # Preserve any panels already in an existing workbook so partial runs accumulate.
    panels = {}
    if OUT_PATH.exists():
        from openpyxl import load_workbook
        try:
            existing = load_workbook(OUT_PATH)
            for name in existing.sheetnames:
                panels[name] = ("__existing__", existing[name])
        except Exception as exc:
            print(f"[warn] could not load existing workbook: {exc}")

    for name, fn in builders.items():
        print(f"[build] {name} ...", flush=True)
        try:
            panels[name] = ("__new__", fn())
            print(f"[ok]    {name}")
        except Exception:
            print(f"[FAIL]  {name}\n{traceback.format_exc()}")

    wb = Workbook()
    wb.remove(wb.active)
    ordered = [n for n in ORDER if n in panels] + [n for n in panels if n not in ORDER]
    for name in ordered:
        kind, payload = panels[name]
        ws = wb.create_sheet(title=name[:31])
        if kind == "__existing__":
            for row in payload.iter_rows(values_only=True):
                ws.append(row)
        else:
            _write_panel(ws, payload)

    wb.save(OUT_PATH)
    print(f"\nSaved {OUT_PATH} with {len(ordered)} sheets: {ordered}")


if __name__ == "__main__":
    main()
